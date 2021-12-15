import csv
import sys
import os
import time
from copy import copy
from datetime import datetime, date
from functools import reduce

from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

import openpyxl
from openpyxl import Workbook
import pandas as pd

csv.field_size_limit(sys.maxsize)


def formatWS(ws):
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="8EA9DB", end_color='8EA9DB', fill_type="solid")

    dims = {}
    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
            alignment_obj = copy(cell.alignment)
            alignment_obj.horizontal = 'center'
            alignment_obj.vertical = 'center'
            cell.alignment = alignment_obj

            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def loadData():
    directory = os.fsencode(os.getcwd() + '/data')
    for subdir in os.listdir(directory):
        if not subdir.decode().startswith('~'):
            filename = os.path.join(directory, subdir).decode()
            if filename.endswith(".xlsx"):
                parseExcel(filename)


# aux function to compare 2 strings without spaces and casing
def bareCompare(str1, str2):
    return stripdown(str1) == stripdown(str2)


def stripdown(str1):
    return "".join(str1.split()).lower()


def removeSuffixesFromGroup(str1):
    return str.strip(str1).split("-")[0]


def cleanNum(num):
    try:
        return float(num)
    except ValueError:
        if '\n' in num:
            stringarr = num.split('\n')
            return float(stringarr[0])


def stackString(strList):
    strList = cleanList(strList)
    return "\n".join(strList)


def cleanList(strList):
    return list(map(lambda x: x.strftime('%d/%m/%Y') if isinstance(x, datetime) else str(x), strList))


def diff_month(d1, d2):
    return (d1.year - d2.year) * 12 + d1.month - d2.month


def formatDateRangeToString(dateRange):
    dateDict = {}
    for i in range(len(dateRange)):
        monthYearKey = dateRange[i].strftime('%b-%Y')
        dateDict.setdefault(monthYearKey, []).append(dateRange[i])

    finalString = []
    # after sorting them to dicts, go over dict and form string
    for key in dateDict:
        dates = dateDict.get(key)
        dates = sorted(dates)
        # reduce all of the keys
        dayStr = reduce(lambda x, y: x + ',' + y, map(lambda x: x.strftime('%d'), dates))
        finalString.append(dayStr + ' ' + str.replace(key, '-', ' '))
    return "\n".join(finalString)




def genVolunteerSheet(wb, dataDict):
    sheetName = '(D) Volunteers'
    header = ['Individual or Group', 'Individual or Group Number', 'Individual/Group name', 'Status',
              'No. of unique volunteers', 'Name of SSA', 'Centre Name', 'Number of SSAs', 'Programme / Activity name',
              'Number of Roles', 'Event Date', 'Number of Sessions', 'Event Time', 'Hours per session',
              'Total Hours Volunteered', 'Did the volunteer plan and coordinate the activity?',
              'Any additional Remarks']

    ws = wb.create_sheet(sheetName)
    ws.title = sheetName
    ws.append(header)

    for key in dataDict:
        dataArray = dataDict.get(key)
        ssaNames = []
        centreNames = []
        programNames = []
        eventDates = []
        eventTime = []
        hoursPerSess = []
        didVolunteerPlan = []
        remarks = []
        roles = 0
        ssaNum = 0
        for i in range(len(dataArray)):
            dataset = dataArray[i]
            indGrp = dataset[5]
            number = dataset[7]
            grpName = dataset[8]
            status = ''
            if dataset[2] not in ssaNames:
                ssaNum += 1
            if dataset[4] not in programNames:
                # roles seems to be how many unique program entries
                roles += 1
            numVolunteer = dataset[9]
            ssaNames.append(dataset[2])
            centreNames.append(dataset[3])
            programNames.append(dataset[4])
            eventDates.append(dataset[10])
            eventTime.append(dataset[12])
            hoursPerSess.append(dataset[13])
            didVolunteerPlan.append(dataset[16])
            remarks.append(dataset[17])

        sessions = len(programNames)
        totalHoursVolunteer = reduce(lambda a, b: cleanNum(a) + cleanNum(b), hoursPerSess)
        entryData = [indGrp, number, grpName, status, numVolunteer, stackString(ssaNames), stackString(centreNames),
                     ssaNum, stackString(programNames), roles, stackString(eventDates), sessions,
                     stackString(eventTime),
                     stackString(hoursPerSess), totalHoursVolunteer, stackString(didVolunteerPlan),
                     stackString(remarks)]
        ws.append(entryData)
    formatWS(ws)
    return wb


def genProgramSheet(wb, dataDict):
    sheetName = '(G) Programme Details'
    header = ['S/N', 'Programme', 'Partners Involved', 'Programme Details (Date and Time)',
              'Number of Beneficiaries/Service Users', 'Number of Volunteers', '', '', 'Number of Volunteering Hours']

    subHeader = ['', '', '', '', '', 'Adhoc', 'Regular', 'Leader', '']

    ws = wb.create_sheet(sheetName)
    ws.title = sheetName
    ws.append(header)
    ws.append(subHeader)
    # header loop
    # there must be a smarter way to do this but... i don't know
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('A1:A2')
    ws.merge_cells('E1:E2')
    ws.merge_cells('F1:H1')
    ws.merge_cells('I1:I2')

    # proceed official data dissection here
    counter = 1
    for key in dataDict:
        dataArray = dataDict.get(key)
        progDates = []
        progTimes = []
        centres = []
        serviceUsers = 0
        adhocCount = 0
        regularCount = 0
        leaderCount = 0
        for i in range(len(dataArray)):
            dataset = dataArray[i]
            progName = dataset[4]
            partnerName = dataset[2] + " & " + dataset[8]
            if dataset[3] not in centres:
                centres.append(dataset[3])
                serviceUsers += dataset[15]
            progDates.append(dataset[10])
            progTimes.append(dataset[12])
            hoursPerSess = dataset[13]
            volunteernum = dataset[9]

        if len(progDates) <= 1:
            adhocCount += volunteernum
        else:
            regularCount += volunteernum

        totalHoursVolunteer = hoursPerSess * len(dataArray) * volunteernum
        progDates = cleanList(progDates)

        progDatetime = []
        for j in range(len(progDates)):
            progDate = progDates[j]
            progTime = progTimes[j]
            datetimeConcat = progDate + ', ' + progTime
            progDatetime.append(datetimeConcat)

        dataEntry = [counter, progName, partnerName, stackString(progDatetime), serviceUsers, adhocCount,
                     regularCount, leaderCount, totalHoursVolunteer]

        ws.append(dataEntry)
        counter += 1
    formatWS(ws)

    return wb

def genPartnerSheet(wb, partnerDict, SSADict):
    sheetName = '(E) Partners and SSAs'
    header = ['S/N', 'Name of Organisation', 'Type of Partner', 'Number of Volunteers']

    ws = wb.create_sheet(sheetName)
    ws.title = sheetName
    ws.append(header)

    counter = 1
    groups = []

    for key in partnerDict:
        dataArray = partnerDict.get(key)
        volunteerCount = 0
        typeOfPartner = ''
        for i in range(len(dataArray)):
            dataset = dataArray[i]
            # extract group number
            groupNumber = dataset[7]
            if groupNumber not in groups:
                typeOfPartner = dataset[6]
                groups.append(groupNumber)
                volunteerCount += int(dataset[9])

        dataEntry = [counter, key, typeOfPartner, volunteerCount]

        ws.append(dataEntry)
        counter += 1
    SSAs = []
    for key in SSADict:
        dataArray = SSADict.get(key)
        volunteerCount = 0
        for i in range(len(dataArray)):
            dataset = dataArray[i]
            SSAName = dataset[2]
            if SSAName not in SSAs:
                SSAs.append(SSAName)
                volunteerCount += int(dataset[9])
        dataEntry = [counter, SSAName, 'SSA', volunteerCount]
        ws.append(dataEntry)
        counter += 1
    formatWS(ws)
    return wb


def genPartnerDetailsSheet(wb, detailsDict):
    sheetName = '(F) Partnership Details'
    header = ['S/N', 'Name of Organisation/Sector', 'Number of Volunteers Deployed to Organisation/Sector',
              'Date of Deployment of Volunteers to Organisation/Sector',
              'Number of Volunteers Active/Volunteering After 6 Months']

    ws = wb.create_sheet(sheetName)
    ws.title = sheetName
    ws.append(header)

    counter = 1
    keyGroup = []
    for key in detailsDict:
        dataArray = detailsDict.get(key)
        namedKey = ''
        volunteerCount = 0
        dateRanges = []
        for i in range(len(dataArray)):
            dataset = dataArray[i]
            namedKey = dataset[2] + " (" + dataset[8] + ")"
            dateRanges.append(dataset[10])
            if namedKey not in keyGroup:
                keyGroup.append(namedKey)
                volunteerCount += int(dataset[9])

        # okay at this point they are all in datetime format, so i can immediately execute readings on them
        minDate = min(dateRanges)
        maxDate = max(dateRanges)
        diff = diff_month(maxDate, minDate)
        dateStr = formatDateRangeToString(dateRanges)
        continuousMonth = 'NA'
        if diff >= 6:
            continuousMonth = 'YES'
        counter += 1
        dataEntry = [counter, namedKey, volunteerCount, dateStr, continuousMonth]
        ws.append(dataEntry)
    formatWS(ws)
    return wb


def parseExcel(filename):
    # I should be able to generate 4 additional sheets after being provided one sheet
    workbook = openpyxl.load_workbook(filename)
    outbook = Workbook()
    ws = workbook.worksheets[0]
    # index ref
    # [0] (Internal/External),[1] (AMK/SK/PG/Others),[2] (Name of SSA),[3] (Centre Name),
    # [4] (Programme / Activity name),[5] (Individual or Group),[6] (Type of Partner),
    # [7] (Individual or Group Number),[8] (Individual/Group name),[9] (No. of unique volunteers),
    # [10] (Event Date),[11] (No. of Sessions),[12] (Event time),[13] (Hours per session),
    # [14] (Total Number of Volunteering Hours),[15] (No. of Service users),
    # [16] (Did the volunteer plan and coordinate the activity?),[17] (Remarks),"
    dataArray = []
    volunteerDict = {}
    programDict = {}
    partnerDict = {}
    SSADict = {}
    detailsDict = {}
    # preloaded
    for index, row in enumerate(ws.iter_rows(min_row=2)):
        dataArray = list(map(lambda x: x.value, row))
        # group by a few criteria, volunteer, partners, and program
        # print(dataArray)

        # make sure to stripdown all keys to prevent missed entries
        groupName = stripdown(dataArray[8])
        volunteerDict.setdefault(groupName, []).append(dataArray)
        programName = stripdown(dataArray[4])
        programDict.setdefault(programName, []).append(dataArray)
        # use column G as unique identifier
        partnerType = stripdown(dataArray[6])
        if partnerType != 'individual':
            partnerDict.setdefault(removeSuffixesFromGroup(dataArray[8]), []).append(dataArray)
            # for this dict, need to set apart using column C + Column I
            detailsKey = stripdown(dataArray[2]) + "-" + groupName
            detailsDict.setdefault(detailsKey, []).append(dataArray)
        SSADict.setdefault(dataArray[2], []).append(dataArray)



    outbook = genVolunteerSheet(outbook, volunteerDict)
    outbook = genPartnerSheet(outbook, partnerDict, SSADict)
    outbook = genPartnerDetailsSheet(outbook, detailsDict)
    outbook = genProgramSheet(outbook, programDict)
    del outbook['Sheet']

    outbook.save('output-data.xlsx')
    workbook.close()
    outbook.close()


loadData()
